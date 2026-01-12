"""
Excel Generator Module

Responsible for:
1. Creating the final Marine Insurance Declaration Excel file
2. Formatting according to NeoAsia standards
3. Organizing data by currency sections (outbound)

Design Decisions:
1. Uses openpyxl for full Excel formatting control
2. Follows existing declaration template format exactly
3. Supports both IN and OUT sheets with proper structure
"""

import io
import logging
from datetime import date
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.shipment import InboundShipment, OutboundShipment, TransportMode
from config.settings import Settings

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generates the Marine Insurance Declaration Excel file.
    
    Output format matches the existing declaration template with:
    - IN sheet: All inbound shipments in chronological order
    - OUT sheet: Outbound shipments grouped by currency
    """
    
    def __init__(self, settings: Settings):
        self.settings = settings
        
        # Styling
        self.header_fill = PatternFill(
            start_color=settings.output.header_bg_color,
            end_color=settings.output.header_bg_color,
            fill_type="solid"
        )
        self.header_font = Font(
            bold=True,
            color=settings.output.header_font_color
        )
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def generate(
        self,
        inbound_shipments: List[InboundShipment],
        outbound_shipments: List[OutboundShipment],
        declaration_period: str
    ) -> io.BytesIO:
        """
        Generate the complete declaration Excel file.
        
        Args:
            inbound_shipments: List of inbound shipment records
            outbound_shipments: List of outbound shipment records
            declaration_period: Period string like "October-25"
            
        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        
        # Create IN sheet
        ws_in = wb.active
        ws_in.title = f"IN {declaration_period}"
        self._create_inbound_sheet(ws_in, inbound_shipments, declaration_period)
        
        # Create OUT sheet
        ws_out = wb.create_sheet(f"OUT {declaration_period}")
        self._create_outbound_sheet(ws_out, outbound_shipments, declaration_period)
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_inbound_sheet(
        self,
        ws,
        shipments: List[InboundShipment],
        period: str
    ):
        """
        Create the IN (Inbound) sheet.
        
        COLUMN PRUNING (Final Output Only):
        - Removed: Description, Rate, Cost
        - Internal data structures unchanged
        """
        
        # Title row
        ws.merge_cells('B1:R1')
        ws['B1'] = f"SCHEDULE OF INCOMING SHIPMENT DECLARATIONS: {period}"
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center')
        
        # Header row 1 (main headers) - PRUNED: Description(col J), Rate(col K), Cost(col O)
        headers_row1 = [
            '', 'ETD DATE', 'BILL OF LADING /', 'Incoterms', 
            'Mode of transportation', 'VESSEL / TRUCK #', 'VOYAGE', '',
            'BRAND', 'FCL / LCL', 'CURR', 
            'VALUE OF GOODS', 'Reference Document No.',
            'Value (SIN)', 'Value (MAL)', 'Value (VIT)', 'Value (Indonesia)', 'Value (PH)'
        ]
        
        # Header row 2 (sub-headers)
        headers_row2 = [
            '', '', 'AIR WAYBILL /', '', '', 'FLIGHT NO', 'FROM', 'TO',
            '', '', '', '', '', '', '', '', '', ''
        ]
        
        # Write headers
        for col, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        for col, header in enumerate(headers_row2, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        # Data rows - start at row 5
        # Column mapping after pruning:
        # B=Date, C=Tracking, D=Incoterms, E=Mode, F=Flight, G=From, H=To
        # I=Brand, J=FCL/LCL, K=Currency, L=Value, M=Reference
        # N=SIN, O=MAL, P=VIT, Q=Indonesia, R=PH
        for row_idx, shipment in enumerate(shipments, 5):
            # Date
            if shipment.etd_date:
                ws.cell(row=row_idx, column=2, value=shipment.etd_date)
                ws.cell(row=row_idx, column=2).number_format = 'YYYY-MM-DD'
            
            # Tracking/AWB
            ws.cell(row=row_idx, column=3, value=shipment.tracking_or_awb)
            
            # Incoterms
            ws.cell(row=row_idx, column=4, value=shipment.incoterms)
            
            # Mode
            ws.cell(row=row_idx, column=5, value=shipment.mode.value if shipment.mode != TransportMode.UNKNOWN else '')
            
            # Flight/Vessel (only for non-COURIER)
            if shipment.mode != TransportMode.COURIER:
                ws.cell(row=row_idx, column=6, value=shipment.flight_vessel)
            
            # From (Origin)
            ws.cell(row=row_idx, column=7, value=shipment.origin_country)
            
            # To (Destination)
            ws.cell(row=row_idx, column=8, value=shipment.destination_country)
            
            # Brand
            ws.cell(row=row_idx, column=9, value=shipment.get_brand_string())
            
            # FCL/LCL (was col 12, now col 10)
            ws.cell(row=row_idx, column=10, value=self.settings.output.default_fcl_lcl)
            
            # Currency (was col 13, now col 11)
            ws.cell(row=row_idx, column=11, value=shipment.currency)
            
            # Value (was col 14, now col 12)
            ws.cell(row=row_idx, column=12, value=shipment.total_value)
            if shipment.total_value:
                ws.cell(row=row_idx, column=12).number_format = '#,##0.00'
            
            # Reference (was col 16, now col 13)
            ws.cell(row=row_idx, column=13, value=shipment.reference)
            
            # Country splits (was cols 17-21, now cols 14-18)
            splits = shipment.country_splits
            ws.cell(row=row_idx, column=14, value=splits.get('SIN'))
            ws.cell(row=row_idx, column=15, value=splits.get('MAL'))
            ws.cell(row=row_idx, column=16, value=splits.get('VIT'))
            ws.cell(row=row_idx, column=17, value=splits.get('Indonesia'))
            ws.cell(row=row_idx, column=18, value=splits.get('PH'))
            
            # Format split columns
            for col in range(14, 19):
                cell = ws.cell(row=row_idx, column=col)
                if cell.value:
                    cell.number_format = '#,##0.00'
        
        # Adjust column widths
        column_widths = {
            'B': 15, 'C': 18, 'D': 10, 'E': 15, 'F': 18, 'G': 12, 'H': 12,
            'I': 12, 'J': 10, 'K': 8, 'L': 15, 'M': 30,
            'N': 12, 'O': 12, 'P': 12, 'Q': 15, 'R': 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_outbound_sheet(
        self,
        ws,
        shipments: List[OutboundShipment],
        period: str
    ):
        """
        Create the OUT (Outbound) sheet with currency sections.
        
        COLUMN PRUNING (Final Output Only):
        - Removed: Rate, Conversion, Cost, Brand
        - Internal data structures unchanged
        """
        
        # Title row
        ws.merge_cells('B1:J1')
        ws['B1'] = f"SCHEDULE OF OUTGOING SHIPMENT DECLARATIONS: {period}"
        ws['B1'].font = Font(bold=True, size=14)
        ws['B1'].alignment = Alignment(horizontal='center')
        
        # Group shipments by currency
        currency_groups: Dict[str, List[OutboundShipment]] = {}
        for shipment in shipments:
            currency = shipment.currency or 'USD'
            if currency not in currency_groups:
                currency_groups[currency] = []
            currency_groups[currency].append(shipment)
        
        # Write sections in order
        current_row = 4
        
        for currency in self.settings.output.outbound_currency_order:
            records = currency_groups.get(currency, [])
            
            # Section header - PRUNED: Brand(col H), Rate(col J), Conversion(col K), Cost(col N)
            headers = [
                '', 'DATE', 'PROFORMA INV / INV', 'VEHICLE NO / FLIGHT NO',
                'Mode of transport', 'FROM', 'TO', 'DESCRIPTION OF GOODS',
                'FCL/LCL', f'VALUE ({currency})'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.thin_border
                cell.alignment = self.center_align
            
            current_row += 1
            
            # Data rows
            # Column mapping after pruning:
            # B=Date, C=Invoice, D=Flight, E=Mode, F=From, G=To
            # H=Description, I=FCL/LCL, J=Value
            for shipment in records:
                # Date
                if shipment.date:
                    ws.cell(row=current_row, column=2, value=shipment.date)
                    ws.cell(row=current_row, column=2).number_format = 'YYYY-MM-DD'
                
                # Invoice
                ws.cell(row=current_row, column=3, value=shipment.invoice_number)
                
                # Flight/Vehicle
                ws.cell(row=current_row, column=4, value=shipment.flight_vehicle)
                
                # Mode
                ws.cell(row=current_row, column=5, value=shipment.mode.value)
                
                # From
                ws.cell(row=current_row, column=6, value=shipment.origin)
                
                # To
                ws.cell(row=current_row, column=7, value=shipment.destination)
                
                # Description (was col 9, now col 8)
                ws.cell(row=current_row, column=8, value=shipment.description)
                
                # FCL/LCL (was col 12, now col 9)
                ws.cell(row=current_row, column=9, value=shipment.fcl_lcl)
                
                # Value (was col 13, now col 10)
                ws.cell(row=current_row, column=10, value=shipment.value)
                if shipment.value:
                    # Format based on currency
                    if currency in ['IDR', 'VND']:
                        ws.cell(row=current_row, column=10).number_format = '#,##0'
                    else:
                        ws.cell(row=current_row, column=10).number_format = '#,##0.00'
                
                current_row += 1
            
            # Total row
            total_value = sum(s.value or 0 for s in records)
            current_row += 1
            ws.cell(row=current_row, column=9, value='TOTAL')
            ws.cell(row=current_row, column=9).font = Font(bold=True)
            ws.cell(row=current_row, column=10, value=total_value)
            ws.cell(row=current_row, column=10).font = Font(bold=True)
            if currency in ['IDR', 'VND']:
                ws.cell(row=current_row, column=10).number_format = '#,##0'
            else:
                ws.cell(row=current_row, column=10).number_format = '#,##0.00'
            
            current_row += 4  # Space between sections
        
        # Adjust column widths
        column_widths = {
            'B': 15, 'C': 18, 'D': 20, 'E': 15, 'F': 12, 'G': 25,
            'H': 35, 'I': 10, 'J': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
